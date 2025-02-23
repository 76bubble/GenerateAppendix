import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side


"""
start_row: 表头属性行
end_row: 表的最后一行
table_total: [表头的下一行, 表最后一行的下一行）
end_this_table: 所有Option结束后出现的Total行（只有一个Option不存在这一行）

ibar 每个program的第一行
iibar 下一个program的第一行（或整个table的最后一行）
iend Total所在行
"""

results = {}

# tables for total--------------------------
path1 = "output_total.xlsx"
df1 = pd.read_excel(path1, engine="openpyxl")

# "valid percent" separate different questions
valid_percent_indices = df1[
    df1.apply(
        lambda row: row.astype(str).str.contains("valid percent", case=False).any(),
        axis=1,
    )
].index

title_list = []

# Iterate each question
for i in range(len(valid_percent_indices)):
    title_row = valid_percent_indices[i] - 1
    title = df1.iloc[title_row, 0]
    title_list.append(title)
    start_row = valid_percent_indices[i]
    if i + 1 < len(valid_percent_indices):
        subset = df1.iloc[start_row : valid_percent_indices[i + 1]]
        first_empty_row = subset[subset.isna().all(axis=1)].index.min()
        if pd.isna(first_empty_row):
            print("err!")
            exit
        end_row = first_empty_row - 1
    else:
        subset = df1.iloc[start_row : len(df1)]
        first_empty_row = subset[subset.isna().all(axis=1)].index.min()
        if pd.isna(first_empty_row):
            end_row = len(df1) - 1
        else:
            end_row = first_empty_row - 1

    # 1=options, 2=numbers, 4=valid percent
    table_total = df1.iloc[(start_row + 1) : end_row + 1, [1, 2, 4]].copy()
    table_total.columns = ["option", "number", "percent"]

    matches = table_total.iloc[:, 0].apply(lambda x: "Total" in str(x))
    end_this_table = np.where(matches)[0]

    # calculate n=？
    if end_this_table.size > 0:
        table_total = table_total[0 : end_this_table[0]]
        sum_num = table_total.iloc[:, 1].astype(int).sum()
    else:
        table_total = table_total[0:1]
        sum_num = table_total.iloc[0:1, 1].astype(int).sum()

    # Get the value col
    table_total["Total"] = table_total.apply(
        lambda row: f"{row['number']} ({round(float(row['percent']),1)})", axis=1
    )
    table_total = table_total.drop(table_total.columns[1:3], axis=1)

    # Insert the first row with 'n='
    table_total.loc[-1] = ["sum", f"n = {sum_num}(%)"]
    table_total.index = table_total.index + 1
    table_total = table_total.sort_index()

    if title not in results:
        results[title] = {}
    results[title]["Total"] = table_total


# tables for programs-------------------------
path2 = "output_program.xlsx"
df2 = pd.read_excel(path2, engine="openpyxl")

# "valid percent" separate different questions
valid_percent_indices2 = df2[
    df2.apply(
        lambda row: row.astype(str).eq("Percent").any(),  # 精确匹配 "Percent"
        axis=1,
    )
].index

# valid_percent_indices2 = df2[
#     df2.apply(
#         lambda row: row.astype(str).str.contains("Valid Percent", case=False).any(),
#         axis=1,
#     )
# ].index

# Iterate each question
for i in range(len(valid_percent_indices2)):
    if not df2.iloc[valid_percent_indices2[i]].astype(str).eq("Valid Percent").any():
        continue
    title_row2 = valid_percent_indices2[i] - 1
    title2 = df2.iloc[title_row2, 0]
    start_row2 = valid_percent_indices2[i]
    if i + 1 < len(valid_percent_indices2):
        subset2 = df2.iloc[start_row2 : valid_percent_indices2[i + 1]]
        first_empty_row2 = subset2[subset2.isna().all(axis=1)].index.min()
        if pd.isna(first_empty_row2):
            print("err!")
            exit
        end_row2 = first_empty_row2 - 1
    else:
        subset2 = df2.iloc[start_row2 : len(df2)]
        first_empty_row2 = subset2[subset2.isna().all(axis=1)].index.min()
        if pd.isna(first_empty_row2):
            end_row2 = len(df2) - 1
        else:
            end_row2 = first_empty_row2 - 1

    # 0=program, 1=valid, 2=options, 3=numbers, 5=valid percent
    table_program = df2.iloc[(start_row2 + 1) : end_row2 + 1, [0, 1, 2, 3, 5]].copy()
    table_program.columns = ["program", "valid", "option", "number", "percent"]

    # separate different program
    bar = []
    programs = []
    cur_program = None
    for j in range(len(table_program)):
        if pd.notna(table_program.iloc[j, 0]):
            bar.append(j)
            cur_program = table_program.iloc[j, 0]
            programs.append(cur_program)
        elif pd.isna(table_program.iloc[j, 0]) and cur_program is not None:
            table_program.iloc[j, 0] = cur_program

    sum_dict = {}
    for k in range(len(bar)):
        # cur program
        ibar = bar[k]
        # nxt program
        iibar = bar[k + 1] if (k < len(bar) - 1) else len(table_program)
        matches2 = table_program.iloc[ibar:iibar, 2].apply(lambda x: "Total" in str(x))
        # cur program end
        end_this_table2 = np.where(matches2)[0]
        # cur program NO valid
        if table_program.iloc[ibar, 1] != "Valid":
            table_program.iloc[ibar:iibar, :] = np.nan
            sum_dict[programs[k]] = "-"
        else:
            # cur program have >1 value
            if end_this_table2.size > 0:
                iend = ibar + end_this_table2[0]
                table_program.iloc[iend:iibar, :] = np.nan
                sum_dict[programs[k]] = (
                    f"n = {table_program.iloc[ibar:iend, 3].dropna().astype(int).sum()}(%)"
                )
            # cur program have 1 value
            else:
                table_program.iloc[(ibar + 1) : iibar, :] = np.nan
                sum_dict[programs[k]] = (
                    f"n = {table_program.iloc[ibar : (ibar + 1), 3].dropna().astype(int).sum()}(%)"
                )
        # print("k=", k)
        # print("programs[k]", programs[k])
        # print("sum_dict[programs[k]]", sum_dict[programs[k]])

    # Get the value col
    table_program["value"] = table_program.apply(
        lambda row: f"{row['number']} ({round(float(row['percent']),1)})", axis=1
    )

    # pivot
    table_program.dropna(inplace=True, subset=["program"])
    table_program = table_program.pivot(
        index="option", columns="program", values="value"
    )

    expected_programs = [
        "Bachelor of  Nursing",
        "Diploma of Licensed Practical Nursing",
    ]
    # "Bachelor of Science in Psychiatric Nursing" is not included
    existing_programs = table_program.columns.tolist()
    missing_programs = [
        prog for prog in expected_programs if prog not in existing_programs
    ]
    for prog in missing_programs:
        table_program[prog] = np.nan
    table_program = table_program[sorted(expected_programs)]

    # show 'option'
    table_program.reset_index(inplace=True)

    # Insert the first row with 'n='
    sums = [sum_dict[key] for key in sorted(sum_dict.keys())]
    sums.insert(0, "sum")

    #    if len(sums) > 3:
    #        sums = [value for value in sums if value != '-']

    print(sums)
    print(table_program.columns)

    table_program.loc[len(table_program)] = sums
    table_program.index = table_program.index + 1
    table_program = table_program.sort_index()

    if title2 not in results:
        results[title2] = {}
    results[title2]["Other"] = table_program

final_results = {}

for title in results:
    if "Total" in results[title] and "Other" in results[title]:
        total_df = results[title]["Total"]
        other_df = results[title]["Other"]
        if "option" in total_df.columns and "option" in other_df.columns:
            merged_df = pd.merge(other_df, total_df, on="option", how="outer")
            final_results[title] = merged_df
        else:
            print(
                f"Warning: 'option' column not found in one of the DataFrames for title: {title}"
            )
    elif "Total" in results[title]:
        final_results[title] = results[title]["Total"]
    elif "Other" in results[title]:
        final_results[title] = results[title]["Other"]
    else:
        print(f"Warning: No 'Total' or 'Other' found for title: {title}")

# for title in final_results:
#    print(f"Title: {title}")
#    print(final_results[title])

with pd.ExcelWriter("final_results.xlsx", engine="openpyxl") as writer:
    combined_df = pd.DataFrame()

    for title, df in final_results.items():
        output_df = pd.DataFrame(columns=range(len(df.columns)))
        output_df.loc[0, 0] = title

        # header
        headers = df.columns.tolist()
        headers[0] = ""
        output_df.loc[1, :] = headers

        # sum
        sum_row = df.iloc[-1].tolist()
        print(sum_row)
        sum_row[0] = ""
        output_df.loc[2, :] = sum_row

        # option
        for i in range(0, len(df) - 1):
            option_row = df.iloc[i].tolist()
            option_row = [item if pd.notna(item) else "-" for item in option_row]
            output_df.loc[3 + i, :] = option_row

        combined_df = pd.concat([combined_df, output_df], ignore_index=True)
        combined_df = pd.concat([combined_df, pd.DataFrame([""])], ignore_index=True)

    # write
    combined_df.to_excel(writer, sheet_name="Results", index=False, header=False)

wb = openpyxl.load_workbook("final_results.xlsx")
ws = wb.active

for row in ws.iter_rows():
    for cell in row:
        cell.font = Font(name="Arial")

fill = PatternFill(
    start_color="1D55A5", end_color="1D55A5", fill_type="solid"
)  # 首行深蓝色
alignment = Alignment(horizontal="left")
font = Font(name="Arial", color="FFFFFF", bold=True, italic=True)
fill_light = PatternFill(
    start_color="DDE9F5", end_color="DDE9F5", fill_type="solid"
)  # 隔行填浅蓝色
start_light = None


for row in ws.iter_rows(min_col=1, max_col=1, min_row=1, max_row=ws.max_row):
    for cell in row:
        if cell.value in title_list:
            ws.merge_cells(
                start_row=cell.row, start_column=1, end_row=cell.row, end_column=4
            )
            merged_cell = ws.cell(row=cell.row, column=1)
            merged_cell.fill = fill
            merged_cell.alignment = alignment
            merged_cell.font = font
"""            start_light = cell.row

if start_row: 
    for row_idx in range(start_row + 1, ws.max_row + 1): 
        ws.fill = fill_light if (row_idx - start_row) % 2 == 1 else None 
        if fill_color:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col).fill = fill_color
"""

for row in range(1, ws.max_row + 1):
    if all(ws.cell(row=row, column=col).value in [None, ""] for col in range(1, 5)):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)


border = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)

last_row = ws.max_row
for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=4):
    for cell in row:
        cell.border = border

wb.save("Appendix.xlsx")

print("write success")
